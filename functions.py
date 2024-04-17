import time
import datetime
import sqlite3
import os
from openpyxl import Workbook
import math
import string
from win32com import client

#　一か月の勤続状況・支払い金額・保険・合計額を出力yotei 2024/4/17

class Db_setting:
    def __init__(self,database_name):
        self.database_name=database_name
        pass
    
    def create(self,table_name):
        if self.database_name not in os.listdir():
            with open(self.database_name,'w') as f:
                pass
        else:
            pass
        conn=sqlite3.connect(self.database_name)
        cursor=conn.cursor()
        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS {table_name}(
                                id TEXT PRIME KEY,
                                name TEXT,
                                start_time REAL,
                                end_time REAL,
                                rest_time REAL,
                                work_time REAL,
                                salary REAL,
                                start_time_formated TEXT,
                                end_time_formated TEXT,
                                work_time_formated REAL,
                                rest_time_formated REAL,
                                week_day INTEGER
            );
            ''' )
        cursor.close()
        conn.commit()
        conn.close()
        self.table_name=table_name

    def add_record(self,attribute,input,user):
        #change to read record,record[2] record[3]
        #if read record return none, insert id then update name and attribute
        #if return list, do nothing 
        #rewrite it today
        conn=sqlite3.connect(self.database_name)
        cursor=conn.cursor()
        id_format='%Y%m%d'
        date=datetime.datetime.now()
        id=date.strftime(id_format)
        record=self.read_record(id,user)
        if record==None:
             cursor.execute(f'''
                INSERT INTO {self.table_name} (id) VALUES ('{id}');
                ''')
             cursor.execute(f'''
                UPDATE {self.table_name} SET name='{user}' WHERE id='{id}'and name is null;
                ''')
             cursor.execute(f'''
                UPDATE {self.table_name} SET {attribute}='{input}' WHERE id='{id}' and name='{user}';
                ''')
        else:
            cursor.execute(f'''
                UPDATE {self.table_name} SET {attribute}='{input}' WHERE id='{id}' and name='{user}';
                ''')
        cursor.execute(f'''
            UPDATE {self.table_name} SET week_day='{date.weekday()}' WHERE id='{id}' and name='{user}';
            ''')
        conn.commit()
        cursor.close()
        conn.close()
    
    def fill_record(self,record):
        print('start to fill in record')
        rest_time=3600
        salary_call=Salary_cal(record[2],record[3],rest_time,1188,190000,1.25,1.35,record[11],None)
        work_time=salary_call.get_working_time()[0]
        salary=salary_call.calculate_salary()
        work_time_formated=f'{math.floor(work_time/3600)}時間{round(((work_time % 3600)/60),1)}分'
        rest_time_formated=f'{math.floor(rest_time/3600)}時間{round(((rest_time % 3600)/60),1)}分'
        conn=sqlite3.connect(self.database_name)
        cursor=conn.cursor()
        cursor.execute(f'''
        UPDATE {self.table_name}
        SET rest_time='{rest_time}',
            work_time='{work_time}',
            salary='{salary}',
            start_time_formated='{salary_call.format_time(record[2])}',
            end_time_formated='{salary_call.format_time(record[3])}',
            work_time_formated='{work_time_formated}',
            rest_time_formated='{rest_time_formated}'
        WHERE id='{record[0]}'and name='{record[1]}';
        ''')
        cursor.close()
        conn.commit()
        conn.close

    def get_ids(self):
        conn=sqlite3.connect(self.database_name)
        cursor=conn.cursor()
        id_format='%Y%m%d'
        date=datetime.datetime.now()
        id=date.strftime(id_format)
        cursor.execute(f'SELECT id FROM {self.table_name};')
        rows=cursor.fetchall()
        ids=[row[0] for row in rows]
        return ids
    
    def get_names(self):
        conn=sqlite3.connect(self.database_name)
        cursor=conn.cursor()
        cursor.execute(f'SELECT name FROM {self.table_name};')
        rows=cursor.fetchall()
        names=[row[0] for row in rows]
        return names
    
    def read_record(self,id,user):
        conn=sqlite3.connect(self.database_name)
        cursor=conn.cursor()
        cursor.execute(f'''
            SELECT * FROM {self.table_name} WHERE id='{id}' and name='{user}';
            ''')
        conn.commit()
        record=cursor.fetchone()
        cursor.close()
        conn.close()
        return list(record) if record else None
    
    def read_all_records(self, user):
        conn=sqlite3.connect(self.database_name)
        cursor=conn.cursor()
        cursor.execute(f'''
            SELECT * FROM {self.table_name} where name='{user}';
            ''')
        conn.commit()
        records=cursor.fetchall()
        cursor.close()
        conn.close()
        return list(records) if records else None
    
    def get_data(self,month,user):
        ids=self.get_ids()
        names=self.get_names()
        print(names)
        data=[]
        work_time_sum=0
        salary_sum=0
        def map_weekdays(number):
            dict_week={
                0:'月曜日',
                1:'火曜日',
                2:'水曜日',
                3:'木曜日',
                4:'金曜日',
                5:'土曜日',
                6:'日曜日'
            }
            week_day_name=dict_week[number]
            return week_day_name

        for id in set(ids):
            for name in set(names):
                if str(id)[0:6]==month and name==user:
                    print(f'id is {id}, name is {name}')
                    record=self.read_record(id,name)
                    print(f'record is {record}')
                    if record!=None:
                        row=[
                            str(record[0])[0:4]+'年'+str(record[0])[4:6]+'月'+str(record[0])[6:8]+'日',
                            record[1],
                            map_weekdays(record[11]),
                            record[7],
                            record[8],
                            record[10],
                            record[9],
                            f'{round(record[6],1)}円'
                        ]
                        work_time_sum=work_time_sum+round(record[5]/60,1)
                        salary_sum=salary_sum+round(record[6],0)
                        data.append(row)
                    else:
                        pass
                else:
                    print('not found')
       
        def to_date_number(data):
            date_number=int(data[0][0:4])*365+int(data[0][5:7])*30+int(data[0][8:10])
            return date_number

        sorted_data=sorted(data,key=to_date_number)
        header=[
            [f'{month}月'],
            [
                '日付',
            '名前',
            '曜日',
            '出勤時間',
            '退勤時間',
            '休憩時間',
            '労働時間',
            '給料',
            ]
        ]
        summary=[[],['合計労働時間',f'{work_time_sum}分'],['合計給与',f'{salary_sum}円']]
        sorted_data=header+sorted_data+summary
        print(sorted_data)
        return sorted_data
    
    def get_start_info(self,user):
        date=datetime.datetime.now()
        id=date.strftime('%Y%m%d')
        print(id)
        if self.read_record(id,user)==None:
            self.have_start_time=False
            print('no no start time')
        else:
            start_time=self.read_record(id,user)[2]
            if start_time!=None:
                self.have_start_time=True
                print('already checked in')
            else:
                print('no start_time')
                self.have_start_time=False
        return self.have_start_time
    
            

class Salary_cal:
    def __init__(self,start,end,rest,hour_salary,month_salary,overtime_rate,holiday_rate,weekday,user):
        self.start=start
        self.end=end
        self.rest=rest
        self.hour_salary=hour_salary
        self.month_salary=month_salary
        self.legal_work_time=3600*8
        self.overtime_rate=overtime_rate
        self.holiday_rate=holiday_rate
        self.weekday=weekday

    def get_time():
        time_format='%Y年%m月%d日%H時%M分%S秒'
        timestamp=time.time()
        datetime_object=datetime.datetime.fromtimestamp(timestamp)
        formated_time=datetime_object.strftime(time_format)
        return [timestamp,formated_time]
    
    def format_time(self,time_stamp):
        time_format='%Y年%m月%d日%H時%M分%S秒'
        datetime_object=datetime.datetime.fromtimestamp(time_stamp)
        formated_time=datetime_object.strftime(time_format)
        return formated_time

    
    def get_working_time(self):
        if self.end-self.start<6*3600:
            working_time=self.end-self.start
        else:
            working_time=self.end-self.start-self.rest
        formated_time=working_time/3600
        return [working_time,formated_time]
    
    def calculate_salary(self): 
        #　月給制と日給制の違い・祝日を書くべき2024/4/17
        # 労働時間の計算
        if self.end-self.start<6*3600:
            working_time=self.end-self.start
        else:
            working_time=self.end-self.start-self.rest
        # 平日の割増賃金含む一日の賃金
        if working_time<=0:
            salary=0
        else:
            if self.weekday<5:
                if working_time>self.legal_work_time:
                    salary=self.hour_salary*self.legal_work_time/3600+(working_time/3600-self.legal_work_time/3600)*self.hour_salary*self.overtime_rate
                #　割増賃金ないの場合の一日の賃金
                elif working_time<=self.legal_work_time:
                    salary=self.hour_salary*working_time/3600
            else:
                salary=self.hour_salary*working_time*self.holiday_rate/3600
        return salary
        

class Write_excel:
    def __init__(self):
        pass
    def create_workbook(self,data,file_name):
        wb=Workbook()
        ws=wb.active
        for row in data:
            ws.append(row)
        columns=string.ascii_uppercase
        for column in columns:
            ws.column_dimensions[column].width=len(ws['D3'].value)*1.5
        wb.save(file_name)

    def write_pdf(self,data,excel_name,pdf_name):
        wb=Workbook()
        ws=wb.active
        for row in data:
            ws.append(row)
        columns=string.ascii_uppercase
        for column in columns:
            ws.column_dimensions[column].width=len(ws['D3'].value)*1.5
        wb.save(excel_name)
        excel=client.Dispatch('Excel.Application')
        cwd=os.getcwd()
        full_path=os.path.join(cwd,excel_name)
        full_path_pdf=os.path.join(cwd,pdf_name)
        sheets=excel.Workbooks.open(full_path)
        work_sheet=sheets.Worksheets[0]
        work_sheet.PageSetup.Zoom = False
        work_sheet.PageSetup.FitToPagesWide = 1
        work_sheet.PageSetup.LeftMargin = 25
        work_sheet.PageSetup.RightMargin = 25
        work_sheet.PageSetup.TopMargin = 50
        work_sheet.PageSetup.BottomMargin = 50
        work_sheet.ExportAsFixedFormat(0,full_path_pdf)
        try:
            os.system('taskkill /f /im excel.exe')
            sheets.close()
            print('closed')
        except Exception:
            pass
        time.sleep(1)
        os.remove(full_path)






if __name__=="__main__":
    db=Db_setting('ndb.db')
    db.create('time_card')
    db.add_record('start_time',1,'1')
    db.add_record('id', 33,'user')
    print(db.read_record(1))
 
        
        